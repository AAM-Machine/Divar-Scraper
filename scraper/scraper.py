def main():

    import argparse
    import time
    import pandas as pd
    import os
    import random
    import re
    from sqlalchemy import (
        create_engine,
        Column,
        Integer,
        String,
        Text,
        UniqueConstraint,
    )
    import os
    from dotenv import load_dotenv
    load_dotenv()
    from sqlalchemy.ext.declarative import declarative_base
    from sqlalchemy.orm import sessionmaker
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from bs4 import BeautifulSoup

    parser = argparse.ArgumentParser(description="Scrape ads from Divar.")
    parser.add_argument(
        "--pages", type=int, default=6, help="Number of pages to scroll."
    )
    parser.add_argument(
        "--excel", action="store_true", help="Generate an Excel file output."
    )
    args = parser.parse_args()

    Base = declarative_base()

    class DivarAd(Base):
        __tablename__ = "divar_ads"
        id = Column(Integer, primary_key=True, autoincrement=True)
        title = Column(String(256))
        date = Column(String(64))
        meter = Column(String(32))
        year = Column(String(32))
        room = Column(String(32))
        total_price = Column(String(64))
        meter_price = Column(String(64))
        floor = Column(String(32))
        amenities = Column(Text)
        description = Column(Text)
        location = Column(String(128))
        images = Column(Text)
        link = Column(String(512), unique=True, nullable=False)
        __table_args__ = (UniqueConstraint("link", name="uq_divar_link"),)

    DB_USER = os.getenv("DB_USER")
    DB_PASS = os.getenv("DB_PASS")
    DB_HOST = os.getenv("DB_HOST")
    DB_PORT = os.getenv("DB_PORT")
    DB_NAME = os.getenv("DB_NAME")

    DATABASE_URL = (
        f"postgresql://{DB_USER}:{DB_PASS}@{DB_HOST}:{DB_PORT}/{DB_NAME}"
    )
    engine = create_engine(DATABASE_URL)
    Session = sessionmaker(bind=engine)
    session = Session()

    Base.metadata.create_all(engine)

    chrome_options = Options()
    chrome_options.add_argument("--disable-notifications")
    chrome_options.add_argument("--disable-popup-blocking")
    chrome_options.add_argument("--window-size=1366,768")

    driver = webdriver.Chrome(options=chrome_options)

    try:
        driver.get("https://divar.ir/s/tehran/buy-apartment?size=-140")
        print("صفحه دیوار با موفقیت باز شد.")
        scroll_count = args.pages
        for i in range(scroll_count):
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
            print(f"اسکرول {i+1}/{scroll_count} انجام شد")
            time.sleep(5)

        page_source = driver.page_source
        soup = BeautifulSoup(page_source, "html.parser")

        ad_cards = soup.select("div.post-card-item-af972")
        if not ad_cards:
            ad_cards = soup.select("div.kt-post-card")
        if not ad_cards:
            ad_cards = soup.select('a[class*="post-card"]')

        print(f"تعداد {len(ad_cards)} آگهی یافت شد.")

        ad_links = []
        prices = []
        for card in ad_cards:
            try:
                link = None
                if hasattr(card, "name") and card.name == "a":
                    link = card.get("href", "")
                else:
                    link_elem = (
                        card.parent
                        if (
                            hasattr(card, "parent")
                            and hasattr(card.parent, "name")
                            and card.parent.name == "a"
                        )
                        else card.find_parent("a")
                    )
                    if link_elem:
                        link = link_elem.get("href", "")
                if link and not link.startswith("http"):
                    link = "https://divar.ir" + link
                if link:
                    ad_links.append(link)
                    price_elem = card.find("div", class_="kt-post-card__description")
                    price = price_elem.text.strip() if price_elem else "نامشخص"
                    prices.append(price)
            except Exception as e:
                print(f"خطا در استخراج لینک یا قیمت آگهی: {str(e)}")

        print(f"تعداد {len(ad_links)} لینک آگهی جمع‌آوری شد.")

        (
            titles,
            dates,
            meters,
            years,
            rooms,
            meter_prices,
            floors,
            amenities,
            descriptions,
            locations,
            links,
            images,
        ) = ([] for _ in range(12))
        total_prices = prices

        for idx, link in enumerate(ad_links):
            try:
                driver.get(link)
                print(f"({idx+1}/{len(ad_links)}) ورود به آگهی: {link}")
                try:
                    WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.TAG_NAME, "body"))
                    )
                except Exception:
                    time.sleep(5)
                page_source = driver.page_source
                soup = BeautifulSoup(page_source, "html.parser")

                title_tag = soup.find("title")
                title_text = title_tag.text.strip() if title_tag else "نامشخص"
                if " - " in title_text:
                    title, date = title_text.rsplit(" - ", 1)
                else:
                    title, date = title_text, "نامشخص"
                titles.append(title.strip())
                dates.append(date.strip())

                meter = year = room = "نامشخص"
                tr_info = soup.find("tr", class_="kt-group-row__data-row")
                if tr_info:
                    tds = tr_info.find_all("td")
                    if len(tds) >= 3:
                        meter = tds[0].text.strip()
                        year = tds[1].text.strip()
                        room = tds[2].text.strip()
                meters.append(meter)
                years.append(year)
                rooms.append(room)

                meter_price = "نامشخص"
                meter_price_divs = soup.find_all(
                    "div", class_="kt-base-row kt-base-row--large kt-unexpandable-row"
                )
                for div in meter_price_divs:
                    title_box = div.find(
                        "div",
                        class_="kt-base-row__start kt-unexpandable-row__title-box",
                    )
                    value_box = div.find(
                        "div", class_="kt-base-row__end kt-unexpandable-row__value-box"
                    )
                    if title_box and value_box:
                        title_p = title_box.find(
                            "p", class_="kt-base-row__title kt-unexpandable-row__title"
                        )
                        value_p = value_box.find("p", class_="kt-base-row__value")
                        if title_p and value_p and "قیمت هر متر" in title_p.text:
                            meter_price = value_p.text.strip()
                            break
                meter_prices.append(meter_price)

                floor = "نامشخص"
                floor_divs = soup.find_all(
                    "div", class_="kt-base-row__end kt-unexpandable-row__value-box"
                )
                for div in floor_divs:
                    p = div.find("p", class_="kt-unexpandable-row__value")
                    if p and (
                        re.search(r"\d+ از \d+", p.text)
                        or re.search(r"طبقه", p.text)
                        or re.match(r"\d+$", p.text.strip())
                    ):
                        floor = p.text.strip()
                        break
                floors.append(floor)

                amenity_list = []
                amenity_table = soup.find_all("table", class_="kt-group-row")
                if len(amenity_table) > 1:
                    amenity_tr = amenity_table[1].find(
                        "tr", class_="kt-group-row__data-row"
                    )
                    if amenity_tr:
                        for td in amenity_tr.find_all("td"):
                            amenity_list.append(td.text.strip())
                amenities.append(", ".join(amenity_list) if amenity_list else "نامشخص")

                desc = "نامشخص"
                desc_div = soup.find(
                    "div", class_="kt-base-row kt-base-row--large kt-description-row"
                )
                if desc_div:
                    p_tag = desc_div.find(
                        "p",
                        class_="kt-description-row__text kt-description-row__text--primary",
                    )
                    if p_tag:
                        desc = p_tag.text.strip()
                descriptions.append(desc)

                location = "نامشخص"
                subtitle = soup.find("div", class_="kt-page-title__subtitle")
                if subtitle:
                    location = subtitle.text.strip()
                locations.append(location)

                img_urls = []
                for img in soup.find_all("img", class_="kt-image-block__image"):
                    src = img.get("src", "")
                    if src and "photo/neda/post" in src:
                        img_urls.append(src)
                images.append(", ".join(img_urls) if img_urls else "نامشخص")

                links.append(link)

                exists = session.query(DivarAd).filter_by(link=link).first()
                if not exists:
                    ad_obj = DivarAd(
                        title=title.strip(),
                        date=date.strip(),
                        meter=meter,
                        year=year,
                        room=room,
                        total_price=(
                            total_prices[idx] if idx < len(total_prices) else "نامشخص"
                        ),
                        meter_price=meter_price,
                        floor=floor,
                        amenities=(
                            ", ".join(amenity_list) if amenity_list else "نامشخص"
                        ),
                        description=desc,
                        location=location,
                        images=(", ".join(img_urls) if img_urls else "نامشخص"),
                        link=link,
                    )
                    session.add(ad_obj)
                    session.commit()
                else:
                    print(f"آگهی با این لینک قبلاً ذخیره شده است: {link}")

                time.sleep(random.uniform(2, 5))

                print(
                    f"عنوان: {title}\nتاریخ: {date}\nمتراژ: {meter}\nسال ساخت: {year}\nاتاق: {room}\nقیمت کل: {total_prices[idx] if idx < len(total_prices) else meter_price}\nقیمت هر متر: {meter_price}\nطبقه: {floor}\nامکانات: {amenity_list}\nتوضیحات: {desc}\nمحله: {location}\nعکس‌ها: {img_urls}\nلینک: {link}\n-------------------"
                )

            except Exception as e:
                print(f"خطا در استخراج اطلاعات آگهی: {str(e)}")
                titles.append("نامشخص")
                dates.append("نامشخص")
                meters.append("نامشخص")
                years.append("نامشخص")
                rooms.append("نامشخص")
                total_prices.append("نامشخص")
                meter_prices.append("نامشخص")
                floors.append("نامشخص")
                amenities.append("نامشخص")
                descriptions.append("نامشخص")
                locations.append("نامشخص")
                images.append("نامشخص")
                links.append(link if link else "نامشخص")

        min_length = min(
            len(titles),
            len(dates),
            len(meters),
            len(years),
            len(rooms),
            len(total_prices),
            len(meter_prices),
            len(floors),
            len(amenities),
            len(descriptions),
            len(locations),
            len(images),
            len(links),
        )
        titles = titles[:min_length]
        dates = dates[:min_length]
        meters = meters[:min_length]
        years = years[:min_length]
        rooms = rooms[:min_length]
        total_prices = total_prices[:min_length]
        meter_prices = meter_prices[:min_length]
        floors = floors[:min_length]
        amenities = amenities[:min_length]
        descriptions = descriptions[:min_length]
        locations = locations[:min_length]
        images = images[:min_length]
        links = links[:min_length]

        data = {
            "عنوان آگهی": titles,
            "تاریخ انتشار": dates,
            "متراژ": meters,
            "سال ساخت": years,
            "تعداد اتاق": rooms,
            "قیمت کل": total_prices,
            "قیمت هر متر": meter_prices,
            "طبقه": floors,
            "امکانات": amenities,
            "توضیحات": descriptions,
            "محله یا موقعیت": locations,
            "عکس‌ها": images,
            "لینک آگهی": links,
        }
        df = pd.DataFrame(data)

        if args.excel:
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter

            excel_path = os.path.join("docs", "divar_ads.xlsx")
            os.makedirs("docs", exist_ok=True)
            df.to_excel(excel_path, index=False)
            wb = load_workbook(excel_path)
            ws = wb.active
            for i, col in enumerate(ws.iter_cols(1, ws.max_column), 1):
                max_length = 0
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[get_column_letter(i)].width = min(
                    max_length + 2, 60
                )
            wb.save(excel_path)
            print(
                f"تعداد {len(titles)} آگهی استخراج و در فایل divar_ads.xlsx ذخیره شد."
            )

    except Exception as e:
        print(f"خطای کلی: {str(e)}")
    finally:
        driver.quit()
        print("مرورگر بسته شد.")


if __name__ == "__main__":
    main()
